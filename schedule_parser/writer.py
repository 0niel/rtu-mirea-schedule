import sqlite3
import datetime
import subprocess
import sys


def install(package):
    """
    Устанавливает пакет
    :param package: название пакета (str)
    :return: код завершения процесса (int) или текст ошибки (str)
    """
    try:
        result = subprocess.check_call([sys.executable, '-m', 'pip', 'install', package])
    except subprocess.CalledProcessError as result:
        return result

    return result


try:
    from openpyxl import *
except ImportError:
    exit_code = install("openpyxl")
    if exit_code == 0:
        from openpyxl import *
    else:
        print("При установке пакета возникла ошибка! {}".format(exit_code))
        exit(0)


class New_to_old_table:

    def __init__(self, path_to_template, path_to_db, outfile, start_date, group_names):
        """
        :param outfile: готовое расписание в старом формате
        :param start_date: дата начала учебного семестра str
        :param group_names: спиаок list() групп
        """
        self.template = path_to_template
        self.path_to_db = path_to_db
        self.outfile = outfile
        self.group_names = group_names
        self.start_date = datetime.datetime.strptime(start_date, "%d.%m.%Y")

    def run(self):
        def str_to_list(string):
            lst = []
            string = string.replace(" ", "")
            string = string.replace("'", "")
            string = string.split(",")
            for i in string:
                try:
                    lst.append(int(i))
                except ValueError:
                    continue
            return lst

        wb = load_workbook(self.template)

        # Создание лиситов с названиями групп
        source = wb['template']
        for title in self.group_names:
            target = wb.copy_worksheet(source)
            target.title = title
        wb.remove(source)

        # Чтение из БД
        connect = sqlite3.connect(self.path_to_db)
        cursor = connect.cursor()

        for one_group in self.group_names:
            one_group = one_group.replace('-', '_')

            # формируем dist_list
            dist_list = {}
            temp_list = []
            dist_index = 1

            cursor.execute("SELECT * FROM {}".format(one_group))
            rows = cursor.fetchall()

            for row in rows:
                if row[3] not in temp_list:
                    dist_list[dist_index] = [row[3], row[6]]
                    dist_index += 1
                    temp_list.append(row[3])
            # print(dist_list)
            one_group = one_group.replace('_', '-')

            # записываем dist_list
            group_sheet = wb[one_group]
            for row, index, dist_data in zip(range(5, 5 + len(dist_list)), dist_list.keys(),
                                             dist_list.values()):
                for col in range(42, 49):
                    if col == 42:
                        group_sheet.cell(column=col, row=row, value=index)
                    elif col == 43:
                        group_sheet.cell(column=col, row=row, value=dist_data[0])
                    elif col == 48:
                        group_sheet.cell(column=col, row=row, value=dist_data[1])

            one_group = one_group.replace('-', '_')

            position = [3, 5]
            date = self.start_date
            number_day = date.strftime("%w")
            date = date - datetime.timedelta(days=int(number_day) - 1)
            for week in range(1, 18 + 1):
                if week % 2 == 0:
                    number_week = 2
                else:
                    number_week = 1

                for day in range(1, 6 + 1):
                    # Записываем дату каждого дня
                    group_sheet.cell(column=position[0], row=position[1] - 1, value=date.strftime("%d.%m.%Y"))
                    if day == 6:
                        date = date + datetime.timedelta(days=2)
                    else:
                        date = date + datetime.timedelta(days=1)
                    for para in range(1, 6 + 1):

                        cursor.execute(
                            """SELECT * FROM {} WHERE day = ? AND para = ? AND week = ?;""".format(one_group),
                            (day, para, number_week))
                        rows = cursor.fetchall()

                        for row in rows:
                            include_lst = str_to_list(row[7])
                            except_lst = str_to_list(row[8])
                            show = False
                            # записываем расписание на основе нового индекса дисциплин
                            if not include_lst and not except_lst:
                                show = True
                            elif week in include_lst:
                                show = True
                            elif week in except_lst:
                                show = False

                            if show:
                                for index, dist_data in zip(dist_list.keys(), dist_list.values()):
                                    if dist_data[0] == row[3]:
                                        group_sheet.cell(column=position[0], row=position[1], value=index)
                                try:
                                    room = int(row[5])
                                except ValueError:
                                    room = row[5]

                                group_sheet.cell(column=position[0] + 1, row=position[1], value=room)

                        if para == 6:
                            position[1] = position[1] + 2
                        else:
                            position[1] = position[1] + 1
                position[0] = position[0] + 2
                position[1] = 5

        wb.save(self.outfile)


if __name__ == "__main__":
    wr = New_to_old_table("parser/template.xlsx", "parser/table.db", "parser/out.xlsx", "09.02.2018", ["БСБО-11-17"])
    wr.run()
